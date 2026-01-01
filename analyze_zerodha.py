import openpyxl
import sys

# Load the Zerodha file
file_path = r"c:\Users\Ajay.Gupt\OneDrive - Reliance Corporate IT Park Limited\Documents\csp\tax_hurdle\configuration\taxpnl-UYM760-2024_2025-Q1-Q4.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    print("=" * 80)
    print("ZERODHA FILE ANALYSIS")
    print("=" * 80)
    
    # Sheet names
    print(f"\nNumber of sheets: {len(wb.sheetnames)}")
    print(f"Sheet names: {wb.sheetnames}")
    
    # Analyze first sheet (Tradewise Exits)
    ws = wb.worksheets[0]
    print(f"\n{'='*80}")
    print(f"SHEET 0: {ws.title}")
    print(f"Dimensions: {ws.dimensions}")
    print('='*80)
    
    # Read first 30 rows
    print(f"\nFirst 30 rows:")
    row_num = 0
    for row in ws.iter_rows(min_row=1, max_row=30, max_col=21, values_only=True):
        row_num += 1
        # Filter non-empty cells
        non_empty = [(i, str(v)[:40]) for i, v in enumerate(row) if v is not None and str(v).strip()]
        if non_empty:
            row_str = " | ".join([f"[{i}]{v}" for i, v in non_empty])
            print(f"Row {row_num}: {row_str}")
    
    # Analyze Sheet 1 "Equity and Non Equity"
    if len(wb.worksheets) > 1:
        ws1 = wb.worksheets[1]
        print(f"\n{'='*80}")
        print(f"SHEET 1: {ws1.title}")
        print(f"Dimensions: {ws1.dimensions}")
        print('='*80)
        
        # Read first 30 rows
        print(f"\nFirst 30 rows:")
        row_num = 0
        for row in ws1.iter_rows(min_row=1, max_row=30, max_col=15, values_only=True):
            row_num += 1
            # Filter non-empty cells
            non_empty = [(i, str(v)[:40]) for i, v in enumerate(row) if v is not None and str(v).strip()]
            if non_empty:
                row_str = " | ".join([f"[{i}]{v}" for i, v in non_empty])
                print(f"Row {row_num}: {row_str}")
    
    wb.close()
    
    wb.close()
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
